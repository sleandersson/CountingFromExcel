from openpyxl import load_workbook
import pandas as pd
from tkinter import Tk, Label, Button, messagebox, filedialog, StringVar, IntVar, Checkbutton
from tkinter import ttk
from tkcalendar import DateEntry
import threading
import time
import logging
from datetime import datetime
from tkinter.scrolledtext import ScrolledText

# Setup logging with timestamps
logging.basicConfig(
    filename='error.log',
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Global variables
global stop_timer, stop_process
stop_timer = False
stop_process = False

# Create Tk root window
root = Tk()
root.title("Excel Testname Counter")

# Initialize checkbox variables
log_csv = IntVar(value=1)  # Default checked
log_xlsx = IntVar(value=0)  # Default unchecked

def select_file1():
    file_path = filedialog.askopenfilename(title="Select the first Excel file")
    file1_var.set(file_path)
    file1_display.delete(1.0, "end")
    file1_display.insert("1.0", file_path)

def select_file2():
    file_path = filedialog.askopenfilename(title="Select the second Excel file (optional)")
    file2_var.set(file_path)
    file2_display.delete(1.0, "end")
    file2_display.insert("1.0", file_path)

def count_testnames():
    progress_bar.start()
    elapsed_time_var.set("Elapsed time: 0.00 seconds")

    global stop_timer, stop_process
    stop_timer = False
    stop_process = False

    threading.Thread(target=update_elapsed_time).start()
    threading.Thread(target=process_files).start()

    count_button.config(state='disabled')
    close_button.config(text="Stop", command=stop_search)

def update_elapsed_time():
    start_time = time.time()
    while not stop_timer:
        elapsed_time = time.time() - start_time
        elapsed_time_var.set(f"Elapsed time: {elapsed_time:.2f} seconds")
        time.sleep(0.1)

def stop_search():
    global stop_process
    stop_process = True

def cleanup_and_exit(stopped=False):
    global stop_timer
    stop_timer = True

    progress_bar.stop()

    count_button.config(state='normal')
    close_button.config(text='Close', command=root.quit)

    if stopped:
        messagebox.showinfo("Process Stopped", "The search process was stopped by the user.")

def write_log_to_csv(start_date, end_date, total_count, adm_count, adjusted_total, testname_counts, search_time):
     # Get current date and time for the log filename
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_filename = f"Log_{current_time}.csv"
    
    with open(log_filename, "w") as f:
        # Write the summary section
        f.write("Summary\n")
        f.write(f"Date Range, {start_date.date()} to {end_date.date()}\n")
        f.write(f"Total Count, {total_count}\n")
        f.write(f"Adm Count, {adm_count}\n")
        f.write(f"Adjusted Total (Excluding 'Adm'), {adjusted_total}\n")
        f.write(f"Search Made, {search_time}\n\n")

        # Write the testname counts section
        f.write("Testname Counts\n")
        f.write("Testname,Count\n")
        testname_counts = testname_counts.astype(int)  # Ensure counts are integers
        for testname, count in testname_counts.items():
            f.write(f"{testname},{count}\n")

def write_log_to_excel(start_date, end_date, total_count, adm_count, adjusted_total, testname_counts, search_time):
    # Get current date and time for the log filename
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_filename = f"Log_{current_time}.xlsx"
    
    log_data = {
        "Date Range": [f"{start_date.date()} to {end_date.date()}"],
        "Total Count": [total_count],
        "Adm Count": [adm_count],
        "Adjusted Total (Excluding 'Adm')": [adjusted_total],
        "Search Made": [search_time]
    }
    testname_df = pd.DataFrame(testname_counts, columns=["Count"]).reset_index()
    testname_df.rename(columns={"index": "Testname"}, inplace=True)

    with pd.ExcelWriter(log_filename) as writer:
        summary_df = pd.DataFrame(log_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        testname_df.to_excel(writer, sheet_name="Testname Counts", index=False)

def process_files():
    try:
        start_time = time.time()

        # Parse date range
        start_date = pd.to_datetime(start_date_entry.get_date().strftime('%Y-%m-%d'))
        end_date = pd.to_datetime(end_date_entry.get_date().strftime('%Y-%m-%d'))

        file1 = file1_var.get()
        file2 = file2_var.get()

        if not file1:
            messagebox.showerror("Error", "Please select at least one Excel file.")
            cleanup_and_exit()
            return

        columns = ['Indatum', 'Testnamn']
        filtered_rows = []

        def process_file(file):
            wb = load_workbook(file, read_only=True)
            sheet = wb.active
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
            col_idx = {col: headers.index(col) for col in columns}

            for row in sheet.iter_rows(min_row=2):  # Skip header row
                if stop_process:
                    cleanup_and_exit(stopped=True)
                    return

                indatum = row[col_idx['Indatum']].value
                testnamn = row[col_idx['Testnamn']].value

                if indatum and testnamn:
                    indatum = pd.to_datetime(indatum, errors='coerce')
                    if start_date <= indatum <= end_date:
                        filtered_rows.append({'Indatum': indatum, 'Testnamn': testnamn})

        # Process first file
        process_file(file1)
        if stop_process:
            return

        # Process second file if provided
        if file2:
            process_file(file2)

        if stop_process:
            cleanup_and_exit(stopped=True)
            return

        # Create a DataFrame from filtered rows
        filtered_df = pd.DataFrame(filtered_rows)

        # Process DataFrame in chunks of 10 rows
        total_count = 0
        adm_count = 0
        testname_counts = pd.Series(dtype=int)
        for i in range(0, len(filtered_df), 10):
            if stop_process:
                cleanup_and_exit(stopped=True)
                return

            chunk = filtered_df.iloc[i:i+10]
            total_count += chunk['Testnamn'].count()
            adm_count += chunk[chunk['Testnamn'] == 'Adm'].shape[0]
            testname_counts = testname_counts.add(chunk['Testnamn'].value_counts(), fill_value=0)

        if stop_process:
            cleanup_and_exit(stopped=True)
            return

        # Ensure testname counts are integers
        testname_counts = testname_counts.astype(int)

        # Adjusted total excluding 'Adm'
        adjusted_total = total_count - adm_count

        # Get search timestamp
        search_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Write logs based on checkbox selections
        if log_csv.get():
            write_log_to_csv(start_date, end_date, total_count, adm_count, adjusted_total, testname_counts, search_time)
        if log_xlsx.get():
            write_log_to_excel(start_date, end_date, total_count, adm_count, adjusted_total, testname_counts, search_time)

        # Complete the process
        cleanup_and_exit()
        elapsed_time = time.time() - start_time
        elapsed_time_var.set(f"Elapsed time: {elapsed_time:.2f} seconds")
        messagebox.showinfo(
            "Results",
            f"Period: {start_date.date()} to {end_date.date()}\n"
            f"Adjusted total (excluding 'Adm'): {adjusted_total}\n\n"
            f"Check Log/Logs for details."
        )

    except Exception as e:
        logging.error("An error occurred during processing", exc_info=True)
        messagebox.showerror("Error", f"An error occurred: {e}")
        cleanup_and_exit()

# Set main window size and center it
window_width = 497
window_height = 350
root.geometry(f"{window_width}x{window_height}")

file1_var = StringVar()
file2_var = StringVar()
elapsed_time_var = StringVar(value="Elapsed time: 0.00 seconds")

Label(root, text="Start Date:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
start_date_entry = DateEntry(root, date_pattern='yyyy-mm-dd')
start_date_entry.grid(row=0, column=1, padx=10, pady=5, sticky="w")

Label(root, text="End Date:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
end_date_entry = DateEntry(root, date_pattern='yyyy-mm-dd')
end_date_entry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

Button(root, text="Select First Excel File", command=select_file1).grid(row=2, column=0, padx=10, pady=5, sticky="w")
file1_display = ScrolledText(root, height=2, width=30, wrap="word")
file1_display.grid(row=2, column=1, padx=10, pady=5, sticky="w")

Button(root, text="Select Second Excel File (optional)", command=select_file2).grid(row=3, column=0, padx=10, pady=5, sticky="w")
file2_display = ScrolledText(root, height=2, width=30, wrap="word")
file2_display.grid(row=3, column=1, padx=10, pady=5, sticky="w")

# Add checkboxes for CSV and XLSX logging
Checkbutton(root, text="Log in .csv", variable=log_csv).grid(row=4, column=0, padx=10, pady=5, sticky="w")
Checkbutton(root, text="Log in .xlsx", variable=log_xlsx).grid(row=5, column=0, padx=10, pady=5, sticky="w")

count_button = Button(root, text="Count Testnames", command=count_testnames)
count_button.grid(row=4, column=0, columnspan=2, padx=10, pady=10)

progress_bar = ttk.Progressbar(root, mode='indeterminate')
progress_bar.grid(row=5, column=0, columnspan=2, padx=10, pady=5)

Label(root, textvariable=elapsed_time_var).grid(row=6, column=0, columnspan=2, padx=10, pady=5)

close_button = Button(root, text="Close", command=root.quit)
close_button.grid(row=8, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()
